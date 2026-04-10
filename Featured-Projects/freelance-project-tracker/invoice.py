from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from pathlib import Path
from typing import List


class InvoiceGenerator:
    """Generates professional Excel invoices."""

    OUTPUT_DIR = Path("invoices")

    def __init__(self, project, client, payments: List):
        self.project = project
        self.client = client
        self.payments = payments

    def generate(self):
        """Generate and save the invoice as Excel file."""
        self.OUTPUT_DIR.mkdir(exist_ok=True)

        # Filter payments for this project only
        project_payments = [p for p in self.payments if p.project_id == self.project.id]

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice"

        # Styles
        title_font = Font(name="Arial", size=16, bold=True, color="1F3864")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="1F3864")
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "FREELANCE INVOICE"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        # Invoice Info
        ws["A3"] = "Invoice #:"
        ws["B3"] = f"INV-{self.project.id:04d}"
        ws["A4"] = "Date:"
        ws["B4"] = datetime.now().strftime("%Y-%m-%d")

        # From
        ws["A6"] = "FROM:"
        ws["A6"].font = Font(bold=True)
        ws["A7"] = "Vashu Jain"
        ws["A8"] = "Dadri, Uttar Pradesh, India"

        # To
        ws["C6"] = "TO:"
        ws["C6"].font = Font(bold=True)
        ws["C7"] = self.client.name
        ws["C8"] = self.client.email or ""
        ws["C9"] = self.client.country or ""

        # Project Table Header
        headers = ["Project Title", "Type", "Details", "Rate", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Project Row
        if self.project.project_type == "Fixed":
            details = "Fixed Price Project"
            rate = "-"
            amount = self.project.fixed_price
        else:
            details = f"{self.project.hours} Hours"
            rate = f"₹{self.project.hourly_rate:,.2f}/hr"
            amount = self.project.calculate_total()

        row_data = [
            self.project.title,
            self.project.project_type,
            details,
            rate,
            f"₹{amount:,.2f}"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=12, column=col, value=value)
            cell.border = thin_border
            cell.alignment = center_align

        # Total Due
        ws["D14"] = "TOTAL DUE"
        ws["D14"].font = Font(bold=True)
        ws["E14"] = f"₹{amount:,.2f}"
        ws["E14"].font = Font(bold=True, color="1F3864")

        # Payments Received
        if project_payments:
            ws["A16"] = "Payments Received:"
            ws["A16"].font = Font(bold=True)

            row = 17
            total_paid = 0.0
            for payment in project_payments:
                ws[f"A{row}"] = payment.payment_date
                ws[f"B{row}"] = f"{payment.currency} {payment.amount:,.2f}"
                total_paid += payment.amount
                row += 1

            # Balance Due
            balance = amount - total_paid
            ws[f"D{row}"] = "BALANCE DUE:"
            ws[f"D{row}"].font = Font(bold=True)
            ws[f"E{row}"] = f"₹{balance:,.2f}"
            ws[f"E{row}"].font = Font(bold=True, color="1F3864")

        # Column Widths
        widths = [35, 12, 18, 15, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        # Save File
        filename = self.OUTPUT_DIR / f"Invoice_{self.client.name.replace(' ', '_')}_{self.project.id}.xlsx"
        wb.save(filename)
        print(f"✅ Invoice generated successfully!")
        print(f"📁 Saved as: {filename}")
        return filename