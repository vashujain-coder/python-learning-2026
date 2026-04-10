import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


def load_and_clean(filepath):
    """Load and clean the raw sales data."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"❌ File not found: {filepath}")

    df = pd.read_csv(filepath)
    
    # Standardize column names
    df.columns = df.columns.str.strip().str.title()
    
    # Required columns check
    required_cols = ['Date', 'Sales Person', 'Region', 'Product', 'Units', 'Price']
    missing = [col for col in required_cols if col not in df.columns]
    if missing:
        raise ValueError(f"Missing columns: {missing}")

    # Clean Date
    df['Date'] = pd.to_datetime(df['Date'], format='mixed', errors='coerce')
    df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
    
    # Clean text columns
    for col in ['Sales Person', 'Region', 'Product']:
        df[col] = df[col].str.strip().str.title().fillna('Unknown')
    
    # Clean numeric columns
    df['Units'] = pd.to_numeric(df['Units'], errors='coerce').fillna(0)
    df['Price'] = pd.to_numeric(df['Price'], errors='coerce').fillna(0)
    
    # Calculate Revenue
    df['Revenue'] = df['Units'] * df['Price']
    
    # Drop invalid rows
    df = df.dropna(subset=['Revenue']).reset_index(drop=True)
    
    print(f"✅ Loaded {len(df):,} records from {filepath}")
    return df


def salesperson_summary(df):
    """Generate summary by Sales Person."""
    summary = df.groupby('Sales Person').agg(
        Total_Revenue=('Revenue', 'sum'),
        Total_Units=('Units', 'sum'),
        Total_Orders=('Revenue', 'count')
    ).round(2).reset_index()

    summary = summary.sort_values('Total_Revenue', ascending=False)
    return summary


def region_summary(df):
    """Find best performing product in each region."""
    region_product = df.groupby(['Region', 'Product'])['Revenue'].sum().reset_index()
    best_per_region = region_product.loc[region_product.groupby('Region')['Revenue'].idxmax()]
    best_per_region = best_per_region.rename(columns={'Revenue': 'Revenue'}).round(2)
    best_per_region = best_per_region.sort_values('Revenue', ascending=False)
    return best_per_region


def overall_summary(df):
    """Create overall summary statistics."""
    summary = pd.DataFrame({
        'Metric': [
            'Total Revenue',
            'Total Units Sold',
            'Total Orders',
            'Average Order Value',
            'Number of Sales Persons',
            'Number of Regions',
            'Number of Products'
        ],
        'Value': [
            f"£{df['Revenue'].sum():,.2f}",
            f"{df['Units'].sum():,}",
            f"{len(df):,}",
            f"£{df['Revenue'].mean():,.2f}",
            len(df['Sales Person'].unique()),
            len(df['Region'].unique()),
            len(df['Product'].unique())
        ]
    })
    return summary


def write_report(cleaned_df, salesperson_df, region_df, summary_df, output_file):
    """Write data to Excel with professional formatting."""
    try:
        if os.path.exists(output_file):
            os.remove(output_file)

        # Write Excel file
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            cleaned_df.to_excel(writer, sheet_name='Cleaned Data', index=False)
            salesperson_df.to_excel(writer, sheet_name='Summary by SalesPerson', index=False)
            region_df.to_excel(writer, sheet_name='Best Product by Region', index=False)
            summary_df.to_excel(writer, sheet_name='Overall Summary', index=False)

        # Load for formatting
        wb = load_workbook(output_file)
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', start_color="1F3864")
        center_align = Alignment(horizontal='center')

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 4

        wb.save(output_file)
        print(f"✅ Report successfully created: {output_file}")

    except PermissionError:
        print("❌ Permission denied. Please close the Excel file if it's open.")
    except Exception as e:
        print(f"❌ Error: {e}")


def main():
    try:
        df = load_and_clean('raw_sales.csv')
        
        df_sales = salesperson_summary(df)
        df_region = region_summary(df)
        df_summary = overall_summary(df)

        write_report(df, df_sales, df_region, df_summary, 'sales_report.xlsx')

    except Exception as e:
        print(f"❌ Program failed: {e}")


if __name__ == "__main__":
    main()