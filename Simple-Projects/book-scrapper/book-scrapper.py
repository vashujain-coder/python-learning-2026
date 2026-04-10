import requests
import csv
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup

BASE_URL = "https://books.toscrape.com/catalogue/page-{}.html"


def scrape_page(page_number):
    """Scrape one page"""
    try:
        r = requests.get(BASE_URL.format(page_number))
        r.raise_for_status()
    except Exception as e:
        print(f"Error on page {page_number}: {e}")
        return []

    soup = BeautifulSoup(r.text, "lxml")
    books = soup.find_all("article", class_="product_pod")
    
    result = []
    for book in books:
        title = book.h3.a['title']
        price = book.find('p', class_="price_color").text.strip()
        rating = book.find('p', class_="star-rating")['class'][1]
        result.append({'Title': title, 'Price': price, 'Rating': rating})
    
    return result


def scrape_all_pages():
    """Scrape all 50 pages"""
    all_books = []
    start = time.time()
    
    for page in range(1, 51):
        print(f"Scraping page {page}/50...", end="\r")
        books = scrape_page(page)
        all_books.extend(books)
        time.sleep(0.5)
    
    elapsed = round(time.time() - start, 2)
    print(f"\n✅ Scraping completed in {elapsed} seconds | Total books: {len(all_books)}")
    return all_books


def save_to_csv(books):
    """Save to CSV"""
    fieldnames = ["Title", "Price", "Rating"]
    try:
        with open("books.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(books)
        print("✅ Saved to books.csv")
    except Exception as e:
        print(f"Error saving CSV: {e}")


def analyse_books(books):
    """Convert to pandas DataFrame and clean data"""
    df = pd.DataFrame(books)

    df['Price'] = df['Price'].str.replace('[^0-9.]', '', regex=True).astype(float)
    df.rename(columns={'Price':'Price(£)'},inplace=True)

    rating_map = {'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5}
    df['Rating'] = df['Rating'].map(rating_map)
    
    return df


def generate_report(df):
    """Generate formatted Excel report"""
    with pd.ExcelWriter('books_report.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="All Books", index=False)

        top_rated = df[df['Rating'] >= 4].sort_values("Rating", ascending=False)
        top_rated.to_excel(writer, sheet_name="Top Rated", index=False)

        cheapest = df.nsmallest(20, 'Price(£)')
        cheapest.to_excel(writer, sheet_name="Cheapest 20", index=False)

    wb = load_workbook('books_report.xlsx')
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill('solid', start_color='1F3864')
            cell.alignment = Alignment(horizontal='center')
        
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save('books_report.xlsx')
    print("✅ Formatted Excel report saved: books_report.xlsx")


if __name__ == "__main__":
    books = scrape_all_pages()
    save_to_csv(books)
    
    df = analyse_books(books)
    generate_report(df)

    print("\n" + "="*50)
    print("📊 ANALYSIS SUMMARY")
    print("="*50)
    print(f"Total books scraped : {len(df)}")
    print(f"Average Price       : £{df['Price(£)'].mean():.2f}")
    print(f"Highest Price       : £{df['Price(£)'].max():.2f}")
    print(f"Cheapest Book       : £{df['Price(£)'].min():.2f}")
    print(f"5-Star Books        : {len(df[df['Rating'] == 5])}")
    print(f"4+ Star Books       : {len(df[df['Rating'] >= 4])}")
    print("="*50)